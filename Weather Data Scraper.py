import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime


def get_weather_data():
    url = "https://weather.com/en-TT/weather/today/l/57da434070e8cbb60665060f8d5d3bcb79305857d11b83ebede9b2af6f7a9730"  # Replace with the actual URL
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Raise an exception for bad responses (e.g., 404 or 500)
    except requests.exceptions.RequestException as e:
        print(f"Failed to fetch data: {e}")
        return None

    soup = BeautifulSoup(response.content, "html.parser")

    # Example: Assuming the temperature is in a div with class 'temperature'
    temperature_element = soup.find(
        "span", class_="TodayDetailsCard--feelsLikeTempValue--2icPt"
    )

    if temperature_element:
        temperature = temperature_element.text.strip()
    else:
        temperature = "N/A"
        print("Temperature not found on the webpage. Check the HTML structure.")

    return temperature


def write_to_excel(temperature):
    excel_file = "weather_data.xlsx"

    try:
        try:
            # Load existing workbook
            workbook = load_workbook(excel_file)
            sheet = workbook.active
        except FileNotFoundError:
            # Create a new workbook with headers if the file doesn't exist
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Temperature"])

        # Append current date and temperature to the sheet
        row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), temperature]
        sheet.append(row)

        workbook.save(excel_file)
        print("Weather data written to Excel successfully.")
    except Exception as e:
        print(f"Failed to write data to Excel: {e}")


if __name__ == "__main__":
    temperature = get_weather_data()

    if temperature is not None:
        print(f"Temperature: {temperature}")
        write_to_excel(temperature)
    else:
        print("Failed to retrieve temperature data.")
